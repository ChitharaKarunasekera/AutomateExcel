from openpyxl import Workbook, load_workbook

#####################################################################
# # Section 1
# wb = Workbook()  # initialize new workbook
# ws = wb.active  # Access sheet. The default sheet
# ws.title = "Data"
#
# # inserting many data at once
# ws.append(['Chithara', 'Is', 'Great', '!'])
# ws.append(['Chithara', 'Is', 'Great', '!'])
# ws.append(['Chithara', 'Is', 'Great', '!'])
# ws.append(['Chithara', 'Is', 'Great', '!'])
# ws.append(['Chithara', 'Is', 'Great', '!'])
# ws.append(['End!'])
#
# wb.save('Chithara.xlsx')  # saving file

#####################################################################
# # Section 2
# wb = load_workbook('Grades.xlsx')  # load workbook
# ws = wb.active
#
# # access row 1 through 10
# for row in range(1, 11):
#     for col in range(0, 4):
#         # chr takes an int and gives the character representation of it
#         char = chr(65 + col)
#         print(ws[char + str(row)].value)


#####################################################################.
# Merging cells
wb = load_workbook('Chithara.xlsx')  # load workbook
ws = wb.active

ws.merge_cells("A1:D12")  # merge
# ws.unmerge_cells("A1:D1")  # unmerge

wb.save('Chithara.xlsx')
