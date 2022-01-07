from openpyxl import Workbook, load_workbook  # creat a new workbook

wb = load_workbook('Grades.xlsx')  # loading the workbook. ONLY works with xlsx

# Accessing sheets
ws = wb.active  # access the active worksheet
print(ws)  # name of worksheet

# Create sheet
wb.create_sheet("Test")
print(wb.sheetnames)  # prints out all sheet names

# Access cell values
print(ws['A2'].value)  # access a cell value
ws['A2'].value = "Chithara"  # Changing a cell value

wb.save('Grades.xlsx')  # save the file. Must save after a change
