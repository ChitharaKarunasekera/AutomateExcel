from openpyxl import Workbook, load_workbook  # creat a new workbook

wb = load_workbook('Grades.xlsx')  # loading the workbook. ONLY works with xlsx

# Accessing sheets
ws = wb.active  # access the active worksheet
print(ws)  # name of worksheet

#Access cell values
print(ws['A2'].value)#access a cell value
ws['A2'].value = "Chithara"#Changeing a cell value
print(ws['A2'].value)#access a cell value

wb.save('Grades.xlsx')#save the file
